#!/usr/bin/env python3
"""Convert OCHS metadata xlsx -> data/manuscripts.json (clean Manuscript records).

Adopts OCHS's own field names where possible (compatibility / drop-in), maps a
readable core, and keeps a `raw` passthrough of every original field so nothing
is lost. Provenance: custodian OCHS, CC BY-NC-SA 4.0, source_url preserved.

Usage: python3 scripts/convert-ochs.py <database_metadata.xlsx> <out.json>
"""
import openpyxl, json, re, sys, unicodedata

src = sys.argv[1]
out = sys.argv[2]

wb = openpyxl.load_workbook(src, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
H = list(rows[0])

def slug_of(name_link):
    m = re.search(r"/manuscripts/([^/\"]+)/", str(name_link or ""))
    return m.group(1) if m else None

def strip_html(s):
    if s is None:
        return None
    return re.sub(r"<[^>]+>", "", str(s)).strip() or None

# readable field mapping: our field -> OCHS column
MAP = {
    "title": "Name (transliteration)",
    "titleIndic": "Name (indic script)",
    "titleTranslation": "Name (translation)",
    "alternateTitles": "Alternative names",
    "author": "Author",
    "language": "Language",
    "script": "Script",
    "patron": "Patron of manuscript",
    "scribe": "Scribe of manuscript",
    "sponsor": "Sponsor for digitised manuscript",
    "provenanceCategory": "Provenance category",
    "provenanceNote": "Provenance note",
    "repository": "Current location of physical manuscript",
    "catalogueIds": "External catalogue ids",
    "externalMetadata": "External meta data",
    "multiText": "Multi-text manuscript",
    "dateOriginal": "Dating of original text",
    "dateCopy": "Date of copying",
    "place": "Place of production",
    "material": "Material",
    "lengthCm": "Length (cm)",
    "heightCm": "Height (cm)",
    "condition": "Condition",
    "binding": "Binding hole(s)",
    "folios": "Folios/pages",
    "linesPerFolio": "Lines per folio/page",
    "foliation": "Foliation/pagination",
    "volumes": "Number of volumes",
    "tradition": "Tradition(s) & subject(s)",
    "transliteration": "Transliteration",
    "translations": "Translation(s)",
    "cataloguers": "Cataloguer(s)",
    "incipit": "Excerpt: Beginning",
    "explicit": "Excerpt: End",
    "subColophon": "Excerpt: Sub-colophon(s)",
    "colophon": "Excerpt: Colophon",
    "secondaryLiterature": "Relevant secondary literature",
    "illustrations": "Illustrations",
    "remarks": "Remarks",
}

records = []
for r in rows[1:]:
    row = dict(zip(H, r))
    slug = slug_of(row.get("Name (transliteration)"))
    if not slug:
        continue
    rec = {
        "id": f"tk:ms:{slug}",
        "ochs_slug": slug,
        "custodian": "OCHS",
        "licence": "CC-BY-NC-SA-4.0",
        "source_url": f"https://ochs-database.netlify.app/manuscripts/{slug}/",
        "photos": str(row.get("Photos available?")).lower() not in ("none", "false", "false"),
        "text": str(row.get("Text available?")).lower() not in ("none", "false", "false"),
    }
    for k, col in MAP.items():
        rec[k] = strip_html(row.get(col))
    # readable title fallback (from the link text)
    if rec["title"] is None:
        rec["title"] = strip_html(row.get("Name (transliteration)") or slug)
    rec["raw"] = {str(k): (None if v is None else str(v)) for k, v in row.items()}
    records.append(rec)

json.dump(records, open(out, "w"), ensure_ascii=False, indent=1)
print(f"wrote {len(records)} manuscript records -> {out}")
